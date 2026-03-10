"""
Fetch real shipping costs from recent sales via ML Orders + Shipments API.
Usage: python scripts/fetch_ml_shipping_costs.py [LOOK_BACK_DAYS] [MLB_ID1 MLB_ID2 ...]
Default: 30 days, items MLB5128815882 MLB4985582952 MLB4553398896 MLB3661932299 MLB3790690639 MLB4252641399
"""
import asyncio
import json
import sys
import os
from datetime import datetime, timedelta, timezone

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import httpx
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from mercadolivre_service import get_valid_access_token
from config import settings

DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql+psycopg://app_ads_generator_usr:app_ads_generator_psw@localhost:5432/app_ads_generator_db",
)
engine = create_engine(DATABASE_URL, future=True)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)

from app import UserConfig

ML_API_BASE = "https://api.mercadolibre.com"
USER_ID = "182bd036-b52b-40ce-b027-b34e0451f64c"
ML_USER_ID = "1452969010"

DEFAULT_ITEMS = [
    "MLB5128815882", "MLB4985582952", "MLB4553398896",
    "MLB3661932299", "MLB3790690639", "MLB4252641399",
]


async def get_token():
    db = SessionLocal()
    try:
        cfg_row = db.query(UserConfig).filter(UserConfig.user_id == USER_ID).first()
        if not cfg_row:
            raise Exception("UserConfig not found")
        ml_accounts = (cfg_row.data or {}).get("ml_accounts") or []
        if not ml_accounts:
            raise Exception("No ML accounts configured")
        account = ml_accounts[0]
    finally:
        db.close()

    access_token, updated = await get_valid_access_token(
        account, settings.ml_client_id, settings.ml_client_secret
    )

    if updated:
        db = SessionLocal()
        try:
            cfg_row = db.query(UserConfig).filter(UserConfig.user_id == USER_ID).first()
            ml_accounts = list((cfg_row.data or {}).get("ml_accounts") or [])
            ml_accounts[0] = updated
            data = dict(cfg_row.data or {})
            data["ml_accounts"] = ml_accounts
            cfg_row.data = data
            db.commit()
            print("[Token refreshed and saved]")
        finally:
            db.close()

    return access_token


async def fetch_orders_for_item(client, headers, item_id, date_from, date_to):
    """Fetch all orders for a given item_id within date range, handling pagination."""
    orders = []
    offset = 0
    limit = 50

    while True:
        resp = await client.get(
            f"{ML_API_BASE}/orders/search",
            params={
                "seller": ML_USER_ID,
                "order.date_created.from": date_from,
                "order.date_created.to": date_to,
                "q": item_id,
                "sort": "date_asc",
                "limit": limit,
                "offset": offset,
            },
            headers=headers,
        )

        if resp.status_code != 200:
            print(f"  [WARN] Orders search failed for {item_id}: {resp.status_code} {resp.text[:200]}")
            break

        data = resp.json()
        results = data.get("results") or []
        orders.extend(results)

        total = data.get("paging", {}).get("total", 0)
        offset += limit
        if offset >= total or not results:
            break

    return orders


async def fetch_shipment_costs(client, headers, shipment_id):
    """Fetch shipping costs for a given shipment_id."""
    resp = await client.get(
        f"{ML_API_BASE}/shipments/{shipment_id}/costs",
        headers={**headers, "x-format-new": "true"},
    )

    if resp.status_code != 200:
        return None

    return resp.json()


async def fetch_item_details(client, headers, item_id):
    """Fetch item to get free_shipping status and list_cost."""
    resp = await client.get(f"{ML_API_BASE}/items/{item_id}", headers=headers)
    if resp.status_code != 200:
        return None, None

    item_data = resp.json()
    shipping = item_data.get("shipping") or {}
    tags = shipping.get("tags") or []
    has_mandatory_free = "mandatory_free_shipping" in tags

    # Get list_cost from free shipping endpoint
    resp2 = await client.get(
        f"{ML_API_BASE}/users/{ML_USER_ID}/shipping_options/free",
        params={"item_id": item_id, "free_shipping": str(has_mandatory_free).lower()},
        headers=headers,
    )

    list_cost = None
    if resp2.status_code == 200:
        coverage = resp2.json().get("coverage", {}).get("all_country", {})
        list_cost = coverage.get("list_cost")

    # Get SKU from attributes
    sku = ""
    for attr in item_data.get("attributes") or []:
        if attr.get("id") == "SELLER_SKU":
            sku = attr.get("value_name", "")
            break

    return sku, list_cost


async def main():
    # Parse args
    args = sys.argv[1:]
    look_back_days = 15
    item_ids = []

    for arg in args:
        if arg.isdigit():
            look_back_days = int(arg)
        else:
            item_ids.append(arg.upper())

    if not item_ids:
        item_ids = DEFAULT_ITEMS

    now = datetime.now(timezone.utc)
    date_to = now.strftime("%Y-%m-%dT%H:%M:%S.000-00:00")
    date_from = (now - timedelta(days=look_back_days)).strftime("%Y-%m-%dT%H:%M:%S.000-00:00")

    print(f"Looking back {look_back_days} days: {date_from} -> {date_to}")
    print(f"Items: {item_ids}\n")

    access_token = await get_token()
    headers = {"Authorization": f"Bearer {access_token}"}

    # Collect results: list of dicts with sku, item_id, list_cost, sender_cost, sender_promoted, receiver_promoted, gross_amount
    results = []

    async with httpx.AsyncClient(timeout=30.0) as client:
        for item_id in item_ids:
            print(f"Processing {item_id}...")

            # Get item details (SKU + list_cost)
            sku, list_cost = await fetch_item_details(client, headers, item_id)
            print(f"  SKU: {sku}, list_cost: {list_cost}")

            # Fetch orders
            orders = await fetch_orders_for_item(client, headers, item_id, date_from, date_to)
            print(f"  Found {len(orders)} orders")

            # Extract shipping IDs
            shipping_ids = set()
            for order in orders:
                ship = order.get("shipping", {})
                ship_id = ship.get("id")
                if ship_id:
                    shipping_ids.add(ship_id)

            print(f"  Unique shipment IDs: {len(shipping_ids)}")

            # Fetch costs for each shipment
            for ship_id in shipping_ids:
                costs = await fetch_shipment_costs(client, headers, ship_id)
                if not costs:
                    print(f"    [WARN] Could not fetch costs for shipment {ship_id}")
                    continue

                gross_amount = costs.get("gross_amount", 0)

                # Sender costs
                sender_cost = 0
                sender_promoted = 0
                for sender in costs.get("senders") or []:
                    sender_cost = sender.get("cost", 0)
                    for disc in sender.get("discounts") or []:
                        sender_promoted += disc.get("promoted_amount", 0)

                # Receiver costs
                receiver_promoted = 0
                receiver = costs.get("receiver") or {}
                for disc in receiver.get("discounts") or []:
                    receiver_promoted += disc.get("promoted_amount", 0)

                results.append({
                    "sku": sku,
                    "item_id": item_id,
                    "shipment_id": ship_id,
                    "list_cost": list_cost,
                    "sender_cost": sender_cost,
                    "sender_promoted": sender_promoted,
                    "receiver_promoted": receiver_promoted,
                    "gross_amount": gross_amount,
                })

    # Print summary
    print(f"\nTotal shipments: {len(results)}")

    # Generate Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from collections import Counter

    wb = Workbook()

    # ── Sheet 1: Resumo (distinct with count) ──
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"

    headers_resumo = ["SKU", "list_cost", "sender_cost", "sender_promoted", "receiver_promoted", "count"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers_resumo, 1):
        cell = ws_resumo.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Group by (sku, list_cost, sender_cost, sender_promoted, receiver_promoted)
    counter = Counter()
    for r in results:
        key = (r["sku"], r["list_cost"] or 0, r["sender_cost"], r["sender_promoted"], r["receiver_promoted"])
        counter[key] += 1

    sorted_groups = sorted(counter.items(), key=lambda x: (x[0][0], -x[1]))
    for row_idx, (key, count) in enumerate(sorted_groups, 2):
        sku, list_cost, sender_cost, sender_promoted, receiver_promoted = key
        values = [sku, list_cost, sender_cost, sender_promoted, receiver_promoted, count]
        for col, val in enumerate(values, 1):
            cell = ws_resumo.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if col >= 2:
                cell.number_format = '#,##0.00' if col < 6 else '#,##0'
                cell.alignment = Alignment(horizontal="right")

    # Auto-width
    for col in range(1, len(headers_resumo) + 1):
        max_len = max(len(str(ws_resumo.cell(row=r, column=col).value or "")) for r in range(1, ws_resumo.max_row + 1))
        ws_resumo.column_dimensions[chr(64 + col)].width = max(max_len + 2, 12)

    # ── Sheet 2: Dados completos ──
    ws_full = wb.create_sheet("Dados Completos")

    headers_full = ["SKU", "item_id", "shipment_id", "list_cost", "sender_cost", "sender_promoted", "receiver_promoted", "gross_amount"]
    for col, h in enumerate(headers_full, 1):
        cell = ws_full.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, r in enumerate(results, 2):
        values = [r["sku"], r["item_id"], r["shipment_id"], r["list_cost"] or 0,
                  r["sender_cost"], r["sender_promoted"], r["receiver_promoted"], r["gross_amount"]]
        for col, val in enumerate(values, 1):
            cell = ws_full.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if col >= 4:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    # Auto-width
    for col in range(1, len(headers_full) + 1):
        col_letter = chr(64 + col) if col <= 26 else chr(64 + (col - 1) // 26) + chr(64 + (col - 1) % 26 + 1)
        max_len = max(len(str(ws_full.cell(row=r, column=col).value or "")) for r in range(1, min(ws_full.max_row + 1, 100)))
        ws_full.column_dimensions[col_letter].width = max(max_len + 2, 14)

    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ml_shipping_costs.xlsx")
    wb.save(output_path)
    print(f"\nExcel saved to: {output_path}")


if __name__ == "__main__":
    asyncio.run(main())
